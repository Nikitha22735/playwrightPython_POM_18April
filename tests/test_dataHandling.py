
import csv
import json
import os

from dotenv import load_dotenv
from openpyxl import load_workbook
import pytest

from utils.jsonhandling import jsonhandling


filePathJson = "testData/credentials.json"


def test_jsonhandling():
    with open(filePathJson) as data:
        creds = json.load(data)
        print(creds["email"])


def test_jsonFromUtils():
    creds = jsonhandling("testData/json2.json")
    print(creds[0]["email"])
    # print(creds["negitiveCreds"]["email"])


def test_jsonFromUtils2():
    print(jsonhandling("testData/credentials.json"))



# =========================================csv===================================
# @pytest.mark.datahandling
def test_handlingCsv():
    finalData =[]
    with open("testData/credentails.csv") as data:
        formattedData = csv.DictReader(data)
        for i in formattedData:
            finalData.append(i)

    print(finalData)


# @pytest.mark.datahandling
def test_writingCsv():
    # finalData =[]
    with open("testData/credentails.csv", mode='w',newline="") as data:
        newData = csv.DictWriter(data,fieldnames=["username","password"])
        newData.writeheader()
        newData.writerow({'username': 'tripur123', 'password': 'welcome123'})


def writongcsv(usname,ps):
     with open("testData/credentails.csv", mode="a",newline="") as data:
        newData = csv.DictWriter(data,fieldnames=["username","password"])
        # newData.writeheader()
        newData.writerow({'username': usname, 'password': ps})



# =========================================excel===================================
#python -m pip install openpyxl

# @pytest.mark.datahandling
def test_handlingExcel():
    values =[]
    filePath_excel= "testData/sample_creds.xlsx"
    workbook = load_workbook(filePath_excel)
    sheet = workbook["Sheet2"]
    for i in sheet.iter_rows(min_row=2, values_only=True):
        values.append(i)

    print(values)


    # sheet.delete_rows(2,1)
    # sheet["A2"] = "test11"
    # sheet["B2"] = "newline11"
    # sheet.append(["test12","newline12"])
    # workbook.save(filePath_excel)


# ====================================cli=============================================
# set usname=Tripura&&set pw=Welcome123&&pytest -m datahandling -s
# $env:usname=Tripura;$env:pw=Welcome123;pytest -m datahandling -s
@pytest.mark.login

def test_cli():
    userName1_1 = os.getenv("usname")
    password1_1 = os.getenv("pw")
    if userName1_1 is None or password1_1 is None:
        print("Environment variables 'usname' and 'pw' are not set.")
    else:
        print("username and password are fetched from cli successfully")

# ===============================env file===========================================
# pip install python-dotenv
# set ENV=dev&&pytest -m datahandling -s
# $env:ENV=dev;pytest -m datahandling -s
# @pytest.mark.datahandling
def test_env():
    # load_dotenv(".env.prod")
    load_dotenv(load_dotenv((f".env.{os.getenv('ENV')}")))
    userName1_1 = os.getenv("usname1")
    password1_1 = os.getenv("pw1")
    url = os.getenv("url1")
    print(userName1_1)
    print(password1_1)
    print(url)
    print("hello")
    

   




    
