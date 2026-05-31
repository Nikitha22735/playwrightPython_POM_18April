from openpyxl import load_workbook


def handlingExcel():
    values =[]
    filePath_excel= "testData/sample_creds.xlsx"
    workbook = load_workbook(filePath_excel)
    sheet = workbook["Sheet2"]
    for i in sheet.iter_rows(min_row=2, values_only=True):
        values.append(i)

    return values